import komm
import numpy as np
import random
import jsonplus
import jsonplus as json
import openpyxl
from os import system, name

def main():
    while True:
        choice = get_main_input()
        if choice == 0:
            exit(0)
        elif choice == 1:
            generate_data()
        elif choice == 2:
            options()
        elif choice == 3:
            automatic_tests2()


# Options
amplitudes = 1.0,  # Nie usuwac "," bo nie bedzie tuple
phase_offsets = 0.0,  # UP
orders = 4,  # UP
snr = 100.0  # Signal to Noise Ratio
signal_power = 1.0  # Moc sygnalu
number_of_tests = 50  # Liczba testow
signal_len = 32  # Dlugosc sygnalu
# digits = 4  # Dokladnosc zaokraglenia


def get_main_input():
    clear()
    print("""========================
M E N U      G L O W N E
========================
1. Generacja danych
2. Ustawienia programu
3. Testy automatyczne
0. Zakoncz
========================""")
    choice = input(">>")
    while choice == "":
        choice = input(">>")

    if int(choice) in [0, 1, 2, 3]:
        return int(choice)

    return get_main_input()


def generate_data():
    global amplitudes, phase_offsets, orders, snr, signal_power, number_of_tests, signal_len #, digits
    sheet_name = input(">> Nazwa arkusza: ")
    # Kanal i modulacja
    channel = komm.AWGNChannel(snr=snr, signal_power=signal_power)
    modulation = komm.APSKModulation(orders=orders, amplitudes=amplitudes, phase_offsets=phase_offsets)
    # Wypisanie wynikow od excela
    wb = openpyxl.load_workbook(filename="Wyniki.xlsx")
    ws = wb.create_sheet(sheet_name)  # Utworzenie arkusza
    # Ustalenie naglowkow
    ws['A1'] = "Numer"
    ws['B1'] = "BER"
    ws['C1'] = "Szybkosc transmisji"
    # Wypisanie danych symulacji
    ws['M1'] = "Amplituda"
    ws['M2'] = "Przesuniecie fazowe"
    ws['M3'] = "Rzad"
    ws['M4'] = "SNR"
    ws['M5'] = "Moc sygnalu"
    ws['M6'] = "Liczba testow"
    ws['M7'] = "Dlugosc sygnalu"
    ws['N1'] = str(amplitudes)
    ws['N2'] = str(phase_offsets)
    ws['N3'] = str(orders)
    ws['N4'] = str(snr)
    ws['N5'] = str(signal_power)
    ws['N6'] = str(number_of_tests)
    ws['N7'] = str(signal_len)
    #Petla
    for i in range(1, number_of_tests+1):
        # Sygnaly
        sygnal_wejsciowy = generate_random_signal()
        sygnal_zakodowany = modulation.modulate(sygnal_wejsciowy)
        sygnal_zaklocony = channel(sygnal_zakodowany)
        sygnal_wyjsciowy = modulation.demodulate(sygnal_zaklocony)
        #BER
        ber = get_ber(sygnal_wejsciowy, sygnal_wyjsciowy)
        ws.cell(row=i+1, column=1).value = i  # Numer
        ws.cell(row=i+1, column=2).value = ber  # BER
        ws.cell(row=i+1, column=3).value = modulation.bits_per_symbol
    wb.save("Wyniki.xlsx")
    print(f"Zakonczono testy - wyniki znajduja sie w pliku Wyniki.xlsx w arkuszu {sheet_name}")
    input("Wcisnij ENTER, aby kontynuowac ...")


def options():
    clear()
    global amplitudes, phase_offsets, number_of_tests, signal_len, orders
    print(f"""========================
USTAWIENIA SYGNALU
------------------------
1. Amplituda = {str(amplitudes)[1:-1]}
2. Przesunięcie = {str(phase_offsets)[1:-1]}
3. Rzad = {str(orders)[1:-1]}
========================
USTAWIENIA KANALU
------------------------
4. SNR = {snr}
5. Moc sygnalu = {signal_power}
========================
USTAWIENIA TESTOW
------------------------
6. Liczba testow = {number_of_tests}
7. Dlugosc sygnalu = {signal_len}
========================
0. Powrot
========================""")
#8. Dokladnosc zaokraglen = {digits}
    choice = int(input(">>"))
    if choice == 0:
        return
    elif choice == 1:
        amplitudes = tuple(map(float, input(">>Amplituda =").split(',')))
    elif choice == 2:
        phase_offsets = tuple(map(float, input(">>Przesuniecie =").split(',')))
    elif choice == 3:
        orders = tuple(map(int, input(">>Rzad =").split(',')))
    elif choice == 4:
        number_of_tests = int(input(">>SNR = : "))
    elif choice == 5:
        number_of_tests = int(input(">>Moc sygnalu = : "))
    elif choice == 6:
        number_of_tests = int(input(">>Liczba testow = : "))
    elif choice == 7:
        signal_len = int(input(">>Dlugosc sygnalu = : "))
    options()


def generate_random_signal(len=signal_len):
    tmp = []
    rng = random.SystemRandom()
    for i in range(len):
        r = rng.random()
        if r < 0.5:
            tmp.append(0)
        else:
            tmp.append(1)
    return tmp


def clear():  #Czyszczenie ekranu
    ## for windows
    #if name == 'nt':
    #    _ = system('cls')
    #    # for mac and linux(here, os.name is 'posix')
    #else:
    #    _ = system('clear')
    print("\n"*5)


def get_ber(s1, s2):
    ber = 0
    for we, wy in zip(s1, s2):
        if we != wy:
            ber += 1
    return ber


def single_test(amplitudes, phase_offsets, orders, snr, signal_power, signal_len, sygnal_wejsciowy):
    channel = komm.AWGNChannel(snr=snr, signal_power=signal_power)
    modulation = komm.APSKModulation(orders=orders, amplitudes=amplitudes, phase_offsets=phase_offsets)
    sygnal_zakodowany = modulation.modulate(sygnal_wejsciowy)
    sygnal_zaklocony = channel(sygnal_zakodowany)
    sygnal_wyjsciowy = modulation.demodulate(sygnal_zaklocony)
    ber = get_ber(sygnal_wejsciowy, sygnal_wyjsciowy)
    return round(ber/signal_len, 3), modulation.bits_per_symbol


def single_test2(channel, modulation, sygnal_wejsciowy):
    sygnal_wyjsciowy = modulation.demodulate(channel(modulation.modulate(sygnal_wejsciowy)))
    ber = get_ber(sygnal_wejsciowy, sygnal_wyjsciowy)
    return round(ber/len(sygnal_wejsciowy), 3)


def automatic_tests():
    with open('test.json', 'r') as read_file:
        cfg = json.loads(read_file.read())
        snr = cfg['snr']
        signal_len = cfg['signal_len']
        signal = generate_random_signal(signal_len)
        signal_power = cfg['signal_power']
        number_of_tests = cfg['number_of_tests']
        print(f"SNR={snr}       Signal_len={signal_len}     Signal_Power={signal_power}     Number_of_tests={number_of_tests}")
        for config in cfg['configs']:
            # Tutaj testy
            print(f"{config['amplitudes']}; {config['phase_offsets']}; {config['orders']};", end=" ")
            for i in range(cfg['number_of_tests']):
                results = single_test(amplitudes=config['amplitudes'], phase_offsets=config['phase_offsets'],
                                      orders=config['orders'], snr=snr, signal_power=signal_power, signal_len=signal_len,
                                      sygnal_wejsciowy=signal)
                if i == 0:
                    print(results[1], end="; # ;")
                print(results[0], end="; ")
            print("")
        input("Waiting for ENTER...")


def automatic_tests2():
    with open('tests_cfg.json', 'r') as read_file:
        cfg = json.loads(read_file.read())
        snr = cfg['snr']
        signal_len = cfg['signal_len']
        signal = generate_random_signal(signal_len)
        signal_power = cfg['signal_power']
        number_of_tests = cfg['number_of_tests']
        print(f"SNR={snr}       Signal_len={signal_len}     Signal_Power={signal_power}     Number_of_tests={number_of_tests}")
        channel = komm.AWGNChannel(snr=snr, signal_power=signal_power)
        for config in cfg['configs']:
            # Tutaj testy
            print(f"{config['amplitudes']}; {config['phase_offsets']}; {config['orders']};", end=" ")
            modulation = komm.APSKModulation(
                orders=config['orders'], amplitudes=config['amplitudes'], phase_offsets=config['phase_offsets'])
            print(modulation.bits_per_symbol, end="; # ;")
            for i in range(cfg['number_of_tests']):
                print(single_test2(channel=channel, modulation=modulation, sygnal_wejsciowy=signal), end="; ")
            print("")
        input("Waiting for ENTER...")


if __name__ == "__main__":
    main()
